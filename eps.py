from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
from lxml import etree
import requests
import pandas as pd
import openpyxl as ope

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=options)




list_dir = r"C:\Users\saeed.moghimzadeh\Desktop\majma.xlsx"
wb_main = ope.load_workbook(list_dir)
ws_main = wb_main.active
list = []
for i in range(2 , ws_main.max_row):
    list.append(ws_main.cell(i , 1).value)



count = 1
data_list = []
for namad in list:
    url_list = []
    url = f"https://www.codal.ir/ReportList.aspx?search&Symbol={namad}&LetterType=6&Isic=271018&AuditorRef=-1&PageNumber=1&Audited&NotAudited&IsNotAudited=false&Childs=false&Mains&Publisher=false&CompanyState=0&Category=1&CompanyType=-1&Consolidatable&NotConsolidatable"
    webpage = driver.get(url)
    time.sleep(1)
    for i in range(1,21):
        try:
            post = driver.find_element(By.XPATH , f'//*[@id="divLetterFormList"]/table/tbody/tr[{i}]/td[8]/div/div[1]/a')
            ref = post.get_attribute("href") + "&sheetid=1"
            url_list.append(ref)
        except:
            continue


    eps = ["error" , "error"]
    sal = ["error" , "error"]
    ids = [3220 , 1464 , 1384 , 1285 , 2149 , 1340 , 2997 , 1145 , 1084]
    for link in url_list:
        webpage = driver.get(link)
        if driver.find_element(By.XPATH , '//*[@id="ctl00_lblPeriod"]').text == '9 ماهه':
            for id in ids:
                try:
                    sal[0] = driver.find_element(By.XPATH , f'//*[@id="{id}"]/thead/tr[1]/th[2]/span').text
                    sal[1] = driver.find_element(By.XPATH , f'//*[@id="{id}"]/thead/tr[1]/th[3]/span').text
                    eps[0] = driver.find_element(By.XPATH , f'//*[@id="{id}"]/tbody/tr[26]/td[2]/span').text
                    eps[1] = driver.find_element(By.XPATH , f'//*[@id="{id}"]/tbody/tr[26]/td[3]/span').text
                    break
                except:
                    continue
            if eps[0] == "error":
                print(namad)
            break
        else:
            continue

    data = {
    "namad" : namad,
    "1سال مالی" : sal[0],
    "eps1" : eps[0],
    "2سال مالی" : sal[1],
    "eps2" : eps[1],
    }
    data_list.append(data)
    print(count , "of " , len(list) , "Done")
    count += 1
    df = pd.DataFrame(data_list)
    df.to_excel("result_dps.xlsx")
