from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
from lxml import etree
import requests
import pandas as pd
import openpyxl as ope

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=options)




list_dir = r"C:\Users\saeed.moghimzadeh\Desktop\majma.xlsx"
wb_main = ope.load_workbook(list_dir)
ws_main = wb_main.active
list = []
for i in range(2 , ws_main.max_row):
    list.append(ws_main.cell(i , 1).value)



count = 1
data_list = []
for namad in list:
    url_list = []
    url = f"https://www.codal.ir/ReportList.aspx?search&Symbol={namad}&LetterType=20&AuditorRef=-1&PageNumber=1&Audited&NotAudited&IsNotAudited=false&Childs&Mains&Publisher=false&CompanyState=-1&Category=-1&CompanyType=-1&Consolidatable&NotConsolidatable"
    webpage = driver.get(url)
    time.sleep(2)
    for i in range(1,21):
        try:
            post = driver.find_element(By.XPATH , f'//*[@id="divLetterFormList"]/table/tbody/tr[{i}]/td[8]/div/div[1]/a')
            ref = post.get_attribute("href") + "&sheetid=1"
            print(ref)
            url_list.append(ref)
        except:
            continue


    dps = ["no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data"]
    eps = ["no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data"]
    sal = ["no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data", "no data"]
    m = 0
    for link in url_list:
        webpage = driver.get(link)
        try:
            sal[m] = driver.find_element(By.XPATH , '//*[@id="lblYearEndToDate"]').text
        except:
            continue
        try:
            eps[m] = driver.find_element(By.XPATH , '//*[@id="ucAssemblyPRetainedEarning_grdAssemblyProportionedRetainedEarning_ctl17_Span1"]').text
        except:
            continue
        try:
            dps[m] = driver.find_element(By.XPATH , '//*[@id="ucAssemblyPRetainedEarning_grdAssemblyProportionedRetainedEarning_ctl18_Span1"]').text
        except:
            continue
        m += 1

    data = {
    "namad" : namad,
    "1سال مالی" : sal[0],
    "dps1" : dps[0],
    "eps1" : eps[0],
    "2سال مالی" : sal[1],
    "dps2" : dps[1],
    "eps2" : eps[1],
    "3سال مالی" : sal[2],
    "dps3" : dps[2],
    "eps3" : eps[2],
    "4سال مالی" : sal[3],
    "dps4" : dps[3],
    "eps4" : eps[3],
    "5سال مالی" : sal[4],
    "dps5" : dps[4],
    "eps5" : eps[4],
    "6سال مالی" : sal[5],
    "dps6" : dps[5],
    "eps6" : eps[5],
    "7سال مالی" : sal[6],
    "dps7" : dps[6],
    "eps7" : eps[6],
    "8سال مالی" : sal[7],
    "dps8" : dps[7],
    "eps8" : eps[7],
    "9سال مالی" : sal[8],
    "dps9" : dps[8],
    "eps9" : eps[8],
    "10سال مالی" : sal[9],
    "dps10" : dps[9],
    "eps10" : eps[9],
    "11سال مالی" : sal[10],
    "dps11" : dps[10],
    "eps11" : eps[10],
    "12سال مالی" : sal[11],
    "dps12" : dps[11],
    "eps12" : eps[11],
    "13سال مالی" : sal[12],
    "dps13" : dps[12],
    "eps13" : eps[13],
    "14سال مالی" : sal[13],
    "dps14" : dps[13],
    "eps14" : eps[13],
    "15سال مالی" : sal[14],
    "dps15" : dps[14],
    "eps15" : eps[14],
    "16سال مالی" : sal[15],
    "dps16" : dps[15],
    "eps16" : eps[15],
    "17سال مالی" : sal[16],
    "dps17" : dps[16],
    "eps17" : eps[16],
    "18سال مالی" : sal[17],
    "dps18" : dps[17],
    "eps18" : eps[17],
    "19سال مالی" : sal[18],
    "dps19" : dps[18],
    "eps19" : eps[18],
    "20سال مالی" : sal[19],
    "dps20" : dps[19],
    "eps20" : eps[19],
    }
    data_list.append(data)
    print(count , "of " , len(list) , "Done" , "  /  " , m , " dps found")
    count += 1
    df = pd.DataFrame(data_list)
    df.to_excel("result_dps.xlsx")
